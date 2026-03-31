from openpyxl import load_workbook
import traceback

def fill_excel_with_template(
    source_path: str = "草稿.xlsx",
    template_path: str = "模板.xlsx",
    output_path: str = "模板填充表.xlsx",
    log_path: str = "error.txt"
):
    """
    根据 source_path(草稿.xlsx) 数据填充 template_path(模板新.xlsx)，
    生成 output_path(模板填充表.xlsx)，并将异常信息记录到 log_path 文件中。
    """

    def log_error(message):
        with open(log_path, "a", encoding="utf-8") as error_file:
            error_file.write(message + "\n")

    try:
        # 加载 Excel 草稿文件
        wb1 = load_workbook(source_path)
        ws1 = wb1.active
        merged_cells = ws1.merged_cells.ranges
        data_dict = {}

        # 处理草稿数据，展开合并单元格
        for row in ws1.iter_rows():
            processed_row = []
            for cell in row:
                if cell.value is None:
                    for merged in merged_cells:
                        min_col, min_row, max_col, max_row = merged.bounds
                        if min_row <= cell.row <= max_row and min_col <= cell.column <= max_col:
                            processed_row.append(ws1.cell(min_row, min_col).value)
                            break
                    else:
                        processed_row.append(None)
                else:
                    processed_row.append(cell.value)

            # 跳过首列为空的行
            if processed_row[0] not in (None, ' '):
                client_name = processed_row[0]
                if client_name not in data_dict:
                    data_dict[client_name] = []
                data_dict[client_name].append(processed_row[1:])

        # 加载模板文件
        wb = load_workbook(template_path)
        ws = wb.active

        # 填充模板
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == '日期':  # 找到日期字段
                    target_row = cell.row - 1
                    target_col = cell.column
                    content = ws.cell(row=target_row, column=target_col).value

                    if content in data_dict:
                        items = data_dict[content]
                        for idx, item_u in enumerate(items, start=1):
                            ws.cell(row=cell.row + idx, column=cell.column + 4, value=item_u[0])
                            ws.cell(row=cell.row + idx, column=cell.column + 5, value=item_u[1])
                            ws.cell(row=cell.row + idx, column=cell.column + 6, value=item_u[2])
                            ws.cell(row=cell.row + idx, column=cell.column + 7, value=item_u[3])

        # 保存文件
        wb.save(output_path)

    except Exception as e:
        log_error(f"发生错误: {str(e)}")
        log_error(f"错误详情: {traceback.format_exc()}")

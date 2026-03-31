from openpyxl import load_workbook
import json
import traceback

def fill_prices_in_excel(
    excel_path: str = "模板填充表.xlsx",
    price_json_path: str = "配置文件/价格表.json",
    diff_json_path: str = "配置文件/差价表.json",
    price_mode_json_path: str = "配置文件/price_mode.json",
    output_path: str = "最终结果.xlsx",
    log_path: str = "error.txt"
):
    """
    从 JSON 价格数据中读取内容，填充到 Excel 中。
    - excel_path: 需要填充价格的 Excel 文件
    - price_json_path: 价格表 JSON 文件
    - diff_json_path: 差价表 JSON 文件
    - price_mode_json_path: 填表模式 JSON 文件
    - output_path: 填充完成后的 Excel 输出文件
    - log_path: 错误日志文件路径
    """

    def log_error(message):
        with open(log_path, "a", encoding="utf-8") as error_file:
            error_file.write(message + "\n")

    def load_json_file(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except FileNotFoundError:
            log_error(f"错误：文件 {json_path} 不存在！")
        except json.JSONDecodeError:
            log_error(f"错误：文件 {json_path} 格式不正确！")
        return {}

    try:
        # 加载 Excel 文件
        wb = load_workbook(excel_path)
        ws = wb.active

        # 加载 JSON 数据
        price_data = load_json_file(price_json_path)
        diff_data = load_json_file(diff_json_path)
        price_mode_data = load_json_file(price_mode_json_path)

        # 遍历所有行，查找“日期”单元格
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == '日期':  # 确保找到"日期"字段
                    target_row = cell.row + 1
                    target_col = cell.column + 2

                    name = ws.cell(row=target_row, column=target_col).value
                    print(f"当前 name: {name}")

                    # 必须三个字典都有该 name
                    if name in price_data and name in diff_data and name in price_mode_data:
                        materials = price_data[name]
                        diff_materials = diff_data[name]
                        price_mode = price_mode_data[name]

                        for i in range(1, 100):
                            if ws.cell(row=cell.row + i, column=5).value is None:
                                break

                            entry = ws.cell(row=cell.row + i, column=cell.column + 5).value

                            if entry in price_mode:
                                val = price_mode[entry]
                                if val == "价格":
                                    ws.cell(row=cell.row + i, column=19, value=materials[entry])
                                elif val == "差价":
                                    ws.cell(row=cell.row + i, column=20, value=diff_materials[entry])
                                elif isinstance(val, dict):  # 是字典才继续判断
                                    if "固定价格" in val:
                                        price = val["固定价格"]
                                        ws.cell(row=cell.row + i, column=19, value=price)
                                    elif "固定差价" in val:
                                        price = val["固定差价"]
                                        ws.cell(row=cell.row + i, column=20, value=price)

        # 保存结果
        wb.save(output_path)

    except Exception as e:
        log_error(f"发生错误: {str(e)}")
        log_error(f"错误详情: {traceback.format_exc()}")
